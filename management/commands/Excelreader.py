import re
import csv
import datetime
from django.conf import settings
from optparse import make_option
from django.core.management.base import BaseCommand, CommandError 
from openpyxl import load_workbook
from openpyxl import Workbook
import arches.app.models.models as archesmodels
from arches.management.commands import utils
from django.core.exceptions import ObjectDoesNotExist


class Command(BaseCommand):
    
    option_list = BaseCommand.option_list + (
        make_option('-o', '--operation', action='store', dest='operation', default='bibliography',type='choice', choices=['bibliography','site_dataset'],help='Operation Type; ' + '\'bibliography\'=Reads a bibliographic XLSX and converts it to arches import file' + '\'site_dataset\'=Reads a site gazetteer XLSX and converts it to arches import file'),
        make_option('-s', '--source', action='store', dest='source', default='',help='Directory containing the XLSX file you need to convert to .arches'),
        make_option('-m', '--mapping', action='store', dest='mapping_file', default='',help='Directory containing a .csv mapping file'),        
        make_option('-d', '--dest_dir', action='store', dest='dest_dir',help='Directory, comprinsing of file name, where you want to save the .arches file'),
        make_option('-r', '--res_type', action='store', dest='resource_type',help='What kind of resource the source file contains e.g. HERITAGE_RESOURCE_GROUP.E27'),
    )
    
    def handle(self, *args, **options):
        print 'operation: '+ options['operation']
        package_name = settings.PACKAGE_NAME
        print 'package: '+ package_name
        
        if options['operation'] == 'site_dataset':
            self.SiteDataset(options['source'], options['resource_type'], options['dest_dir'])
    
    
    def validatedates(self, date):
        try:
            datetime.datetime.strptime(date, '%Y-%m-%d') #Checks for format  YYYY-MM-DD
        except ValueError:
            try:
                d =datetime.datetime.strptime(date, '%Y-%m-%d %X') #Checks for format  YYYY-MM-DD hh:mm:ss
                date = d.strftime('%Y-%m-%d')
            except ValueError:
                try:
                    d = datetime.datetime.strptime(date,'%d-%m-%Y') #Checks for format  DD-MM-YYYY
                    date = d.strftime('%Y-%m-%d')
                except ValueError:
                    try:
                        d = datetime.datetime.strptime(date,'%d/%m/%Y') #Checks for format  DD/MM/YYYY
                        date = d.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            d = datetime.datetime.strptime(date,'%d/%m/%y') #Checks for format  DD/MM/YY
                            date = d.strftime('%Y-%m-%d')
                            
                        except ValueError:
                            try:
                                d = datetime.datetime.strptime(date,'%Y') #Checks for format  YYYY
                                isodate = d.isoformat()
                                date = isodate.strip().split("T")[0] #
                            except:
                                raise ValueError("The value %s inserted is not a date" % date)
        return date
    
#     def validaterows(self,workbook):
    #def validate_resourcetype(self, resourcetype):
        
    def validate_headers(self, workbook):
        for sheet in workbook.worksheets:
            for header in sheet.iter_cols(max_row = 1):
                if header[0].value is not None:
                    try:
                        modelinstance = archesmodels.EntityTypes.objects.get(pk = header[0].value)
                    except ObjectDoesNotExist:
                        raise ObjectDoesNotExist("The header %s is not a valid EAMENA node name" % header[0].value)
        return
#     def validate_concept(self, concept):

    def collect_concepts(self, node_conceptid, full_concept_list = []):
        ''' Collects a full list of child concepts given  the conceptid of the node. Returns a list of a set of concepts, i.e. expounding the duplicates'''
        concepts_in_node = archesmodels.ConceptRelations.objects.filter(conceptidfrom = node_conceptid)
        if concepts_in_node.count() > 0:
            full_concept_list.append(node_conceptid) 
            for concept_in_node in concepts_in_node:
                
                self.collect_concepts(concept_in_node.conceptidto_id, full_concept_list)
        else:
            full_concept_list.append(node_conceptid) 
        return list(set(full_concept_list)) 
    
    def SiteDataset(self, source, resourcetype, destination):
        wb2 = load_workbook(source)
#             self.validaterows(wb2)

        ResourceList = []
        self.validate_headers(wb2)
        for sheet_index,sheet in enumerate(wb2.worksheets):
            sheet_name = wb2.sheetnames[sheet_index]
            for col_index,header in enumerate(sheet.iter_cols(max_row = 1)):
                GroupNo = 0
                if header[0].value is not None:
                    modelinstance = archesmodels.EntityTypes.objects.get(pk = header[0].value)
                    for row_index, row in enumerate(sheet.iter_rows(row_offset = 1)):
    #                     print "Row %s column %s with value %s" %(row_index, col_index, row[col_index].value)
                        if row[col_index].value is not None:
                            value_encoded = (unicode(row[col_index].value)).encode('utf-8') #Values could be in Arabic or containing unicode chars, so it is essential to encode them properly.
                            for concept_index,concept in enumerate(re.sub(ur';\s+', ';', value_encoded).split(';')): #replacing a semicolon (u003b) and space (u0020) with a semicolon in case that that space in front of the semicolon exists
                                GroupName = " ".join((sheet_name, str(GroupNo))) if sheet_name != 'NOT' else sheet_name
                                GroupNo = GroupNo +1 if sheet_name is not 'NOT' else ''
                                                        
                                if modelinstance.businesstablename == 'domains':                                
                                    concepts_in_node = self.collect_concepts(modelinstance.conceptid) 
                                    valuelist = [archesmodels.Values.objects.filter(value = concept, conceptid= concept_in_node) for concept_in_node in concepts_in_node]
                                    valueinstance = [x for x in valuelist if x] 
                                    if valueinstance is not None:
                                        conceptinstance = archesmodels.Concepts.objects.get(pk=valueinstance[0][0].conceptid)
                                        concept_list = [str(row_index),resourcetype,modelinstance.entitytypeid,conceptinstance.legacyoid,GroupName]
                                        ResourceList.append(concept_list)
    #                                     print "ResourceId %s, AttributeName %s, AttributeValue %s, GroupId %s" %(row_index,modelinstance.entitytypeid,conceptinstance.legacyoid,GroupName)
                                if modelinstance.businesstablename == 'strings':
                                        concept_list = [str(row_index),resourcetype,modelinstance.entitytypeid,concept, GroupName]
                                        ResourceList.append(concept_list)
    #                                     print "ResourceId %s, AttributeName %s, AttributeValue %s, GroupId %s" %(row_index,modelinstance.entitytypeid,concept, GroupName)
                                if modelinstance.businesstablename == 'dates':
                                        date = self.validatedates(concept)
                                        concept_list = [str(row_index),resourcetype,modelinstance.entitytypeid,date, GroupName]
                                        ResourceList.append(concept_list)
                                                                        
        with open(destination, 'w') as csvfile:
                writer = csv.writer(csvfile, delimiter ="|")
                writer.writerow(['RESOURCEID', 'RESOURCETYPE', 'ATTRIBUTENAME', 'ATTRIBUTEVALUE', 'GROUPID'])
                for row in ResourceList:
                    writer.writerow(row)


                                    